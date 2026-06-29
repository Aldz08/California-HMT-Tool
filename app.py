# pip install selenium requests pdfplumber pandas openpyxl ttkbootstrap win10toast

import os
import re
import time
import tempfile
import tkinter as tk
from pathlib import Path

import webbrowser
import subprocess
import threading
import requests
import pdfplumber
import pandas as pd

from tkinter import filedialog
from datetime import datetime

import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.widgets import ToastNotification
from ttkbootstrap.dialogs import Messagebox

from winotify import Notification

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC





URL = "https://supreme.courts.ca.gov/case-information/weekly-conference-results"

CCC_URLS = {
    "advance": (
        "https://advance.lexis.com/practice?"
        "config=00JAA3ZTliMzQ2ZC1hM2NkLTRkMDItYThhYS1iN2JmYmU5ZDA5YTQKAFBvZENhdGFsb2dcUEOCJVCGKggHVsu3jpSV"
        "&crid=6c62e7df-f413-43d1-ad65-df7f86dc2bd0"
        "&pdsf=MTA2Mjg2OA~%5Eadministrative-materials~%5ECA+-+California+Compensation+Cases"
        "&pdcontenttypeid="
        "&pdbcts=1780028105811"
        "&prid=c85d0a01-e0ff-47d2-ac84-9308a8cd2e10"
    ),

    "plus": (
        "https://plus.lexis.com/practice/"
        "?config=00JAA3ZTliMzQ2ZC1hM2NkLTRkMDItYThhYS1iN2JmYmU5ZDA5YTQKAFBvZENhdGFsb2dcUEOCJVCGKggHVsu3jpSV"
        "&crid=cffb3cc8-9691-4809-b552-a2149cc6558a"
        "&pdsf=MTA2Mjg2OA~%5Eadministrative-materials~%5ECA%20-%20California%20Compensation%20Cases"
        "&pdcontenttypeid="
        "&pdbcts=1781596075943"
        "&prid=022f8f5e-e28a-4345-91ce-3d4cfe6964c5"
    )
}

SEARCH_BAR_XPATH = '//*[@id="searchTerms"]'
RESULT_ITEMS_CSS = "ol li[data-id^='sr']"

RESULT_LINK_CSS_CANDIDATES = [
    "h2 a",
    "h3 a",
    "a[href*='/document/']",
    "a[href*='document']",
    "a",
]

RESULT_METADATA_CSS_CANDIDATES = [
    "div.metadata span",
    ".metadata span",
    "span",
]

LEXIS_VERSION = "advance"

LEXIS_CONFIGS = {
    "advance": {
        "login_url": (
            "https://signin.lexisnexis.com/lnaccess/app/signin"
            "?back=https%3A%2F%2Fadvance.lexis.com%3A443%2F&aci=la"
        ),
        "result_items_css": "ol li[data-id^='sr']",
        "lexis_cite_css": None,
        "toggle_btn_css": None,
        "hits_css": None,
    },
    "plus": {
        "login_url": (
            "https://signin.lexisnexis.com/lnaccess/app/signin"
            "?back=https%3A%2F%2Fplus.lexis.com%3A443%2Fzhome&aci=lp"
        ),
        "result_items_css": "div.resultsListContainer.row",
        "lexis_cite_xpath": './/results-list-card-metadata-regionalized-host/results-list-card-metadata-default/div/div/span[4]',
        "toggle_btn_css": ".runnewsearchbutton",
        "hits_css": "span.resultNumber",
    },
}

downloaded_pdf_path = None
current_filtered_df = None
current_excel_path = None
imported_pdf_file = False
save_logs_var = None
current_conference_date = ""
username = os.environ.get("USERNAME", "")

def get_selected_lexis_version():
    try:
        return "plus" if lexis_plus_var.get() else "advance"
    except Exception:
        return "advance"

def start_load_links_in_background():
    loading = tb.Toplevel(root)
    loading.title("Loading Links")
    loading.resizable(False, False)

    center_to_parent(
        loading,
        root,
        420,
        200
    )

    loading_label = tb.Label(
        loading,
        text="Loading conference links...",
        font=("Arial", 11, "bold")
    )
    loading_label.pack(pady=(25, 10))

    bar = tb.Progressbar(
        loading,
        mode="indeterminate",
        bootstyle="info-striped",
        length=320
    )
    bar.pack(pady=10)
    bar.start(10)

    load_button.config(state="disabled")

    def cleanup():
        try:
            if bar.winfo_exists():
                bar.stop()
        except tk.TclError:
            pass

        try:
            if loading.winfo_exists():
                loading.destroy()
        except tk.TclError:
            pass

        try:
            if load_button.winfo_exists():
                load_button.config(state="normal")
        except tk.TclError:
            pass

    def worker():
        try:
            links = get_conference_links()

            def success_ui():
                for item in tree.get_children():
                    tree.delete(item)

                if not links:
                    Messagebox.show_warning(
                        "No conference links were found.",
                        "No Links Found"
                    )
                    cleanup()
                    return

                for date_text, href in links:
                    tree.insert(
                        "",
                        END,
                        values=(date_text, href)
                    )

                cleanup()

                show_windows_toast(
                    "Links Loaded",
                    f"Loaded {len(links)} conference link(s)."
                )

            root.after(0, success_ui)

        except Exception as e:
            error_message = str(e)

            def error_ui():
                cleanup()
                Messagebox.show_error(error_message, "Error")

            root.after(0, error_ui)

    threading.Thread(target=worker, daemon=True).start()

def open_conference_website():
    webbrowser.open(
        "https://supreme.courts.ca.gov/case-information/weekly-conference-results"
    )

def open_downloaded_pdf():
    if not downloaded_pdf_path or not os.path.exists(downloaded_pdf_path):
        Messagebox.show_warning(
            message="No downloaded PDF is available to view.",
            title="No PDF"
        )
        return

    try:
        os.startfile(downloaded_pdf_path)
    except Exception as e:
        Messagebox.show_error(
            message=str(e),
            title="Open PDF Error"
        )

def center_to_parent(child, parent, width, height):
    parent.update_idletasks()

    x = parent.winfo_x() + (parent.winfo_width() // 2) - (width // 2)
    y = parent.winfo_y() + (parent.winfo_height() // 2) - (height // 2)

    child.geometry(f"{width}x{height}+{x}+{y}")

def show_windows_toast(title, message):
    toast = Notification(
        app_id="California Supreme Court Tool",
        title=title,
        msg=message
    )

    toast.show()

def show_toast(title, message, style="success"):
    toast = ToastNotification(
        title=title,
        message=message,
        duration=5000,
        bootstyle=style,
        position=(20, 20, "se")
    )
    toast.show_toast()


def normalize_text(value):
    return (
        str(value)
        .replace("–", "-")
        .replace("—", "-")
        .replace("*", "")
        .lower()
        .strip()
    )


def line_text(words):
    return " ".join(
        w["text"] for w in sorted(words, key=lambda x: x["x0"])
    ).strip()


def extract_table_by_positions(pdf_path):
    rows = []

    columns = {
        "Title": (0, 275),
        "Case #": (275, 375),
        "CA #": (375, 490),
        "Action Type": (490, 640),
        "Result": (640, 900),
    }

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=2, y_tolerance=3)

            header_tops = [
                w["top"]
                for w in words
                if w["text"].lower() == "title"
            ]

            if not header_tops:
                continue

            table_start = min(header_tops) + 5
            table_words = [w for w in words if w["top"] > table_start]

            lines = {}

            for w in table_words:
                y = round(w["top"] / 3) * 3
                lines.setdefault(y, []).append(w)

            for y in sorted(lines):
                line_words = lines[y]
                full_line = line_text(line_words)
                lower_line = full_line.lower()

                if full_line.isdigit():
                    continue

                if "supreme court of california" in lower_line:
                    continue

                if "results from the petition conference" in lower_line:
                    continue

                if any(header in lower_line for header in [
                    "title",
                    "case #",
                    "ca#",
                    "ca #",
                    "action type",
                    "result",
                ]):
                    continue

                row = {col: "" for col in columns}

                for col, (x_min, x_max) in columns.items():
                    col_words = [
                        w for w in line_words
                        if x_min <= w["x0"] < x_max
                    ]
                    row[col] = line_text(col_words)

                is_new_case = bool(re.search(r"S\d{6}", row["Case #"]))

                if is_new_case:
                    rows.append(row)
                elif rows:
                    for col in columns:
                        if row[col]:
                            rows[-1][col] = (
                                rows[-1][col] + " " + row[col]
                            ).strip()

    return pd.DataFrame(rows)


def get_conference_links():
    options = webdriver.ChromeOptions()
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--headless=new")

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()

    wait = WebDriverWait(driver, 30)
    results = []

    try:
        driver.get(URL)

        wait.until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "main a")
            )
        )

        time.sleep(1)

        links = driver.find_elements(By.CSS_SELECTOR, "main a")

        for link in links:
            text = link.text.strip()
            href = link.get_attribute("href")

            if not text or not href:
                continue

            if "," in text and any(char.isdigit() for char in text):
                results.append((text, href))

    finally:
        driver.quit()

    return results


def login_lexis(driver, progress_callback=None):
    username = os.getenv("LexisAdvanceUser")
    password = os.getenv("LexisAdvancePass")

    if not username or not password:
        raise RuntimeError(
            "Missing LexisAdvanceUser or LexisAdvancePass environment variable."
        )
    
    wait = WebDriverWait(driver, 30)

    version = get_selected_lexis_version()
    config = LEXIS_CONFIGS[version]

    driver.get(config["login_url"])

    message = f"Logging in to Lexis {version.title()}"
    print(message)
    
    if progress_callback:
        progress_callback(message)

    user_input = wait.until(
        EC.element_to_be_clickable((By.ID, "userid"))
    )
    user_input.clear()
    user_input.send_keys(username)
    user_input.send_keys(Keys.ENTER)

    pass_input = wait.until(
        EC.element_to_be_clickable((By.ID, "password"))
    )
    pass_input.clear()
    pass_input.send_keys(password)
    pass_input.send_keys(Keys.ENTER)

    if version == "advance":
        wait.until(EC.url_contains("advance.lexis.com"))
    else:
        wait.until(
        lambda d:
        "plus.lexis.com" in d.current_url
        or "zhome" in d.current_url
    )

    print("Lexis login successful.")

    return wait


def get_result_cards(driver):
    version = get_selected_lexis_version()
    config = LEXIS_CONFIGS[version]
    return driver.find_elements(By.CSS_SELECTOR, config["result_items_css"])


def get_first_href_from_card(card):
    version = get_selected_lexis_version()

    if version == "plus":
        href = card.get_attribute("href")
        if href:
            return href

        links = card.find_elements(By.CSS_SELECTOR, "a")
        for link in links:
            href = link.get_attribute("href")
            if href:
                return href

        return ""

    for selector in RESULT_LINK_CSS_CANDIDATES:
        links = card.find_elements(By.CSS_SELECTOR, selector)

        for link in links:
            href = link.get_attribute("href")
            if href:
                return href

    return ""


def get_lexis_cite_from_card(card):
    version = get_selected_lexis_version()
    config = LEXIS_CONFIGS[version]

    if version == "plus":
        try:
            element = card.find_element(
                By.XPATH,
                config["lexis_cite_xpath"]
            )

            text = element.text.strip()

            if text:
                return text

        except Exception:
            pass

    time.sleep(0.5)

    card_text = card.text.strip()

    patterns = [
        r"\d{4}\s+[A-Z][A-Za-z. ]+\s+LEXIS\s+\d+",
        r"\d{4}\s+[A-Z][A-Z. ]+\s+LEXIS\s+\d+",
        r"\d{4}\s+LEXIS\s+\d+",
        r"\d+\s+[A-Z][A-Za-z. ]+\s+LEXIS\s+\d+",
    ]

    for pattern in patterns:
        match = re.search(pattern, card_text, flags=re.IGNORECASE)
        if match:
            return match.group(0).strip()

    for selector in RESULT_METADATA_CSS_CANDIDATES:
        elements = card.find_elements(By.CSS_SELECTOR, selector)

        for element in elements:
            text = element.text.strip()

            for pattern in patterns:
                match = re.search(pattern, text, flags=re.IGNORECASE)
                if match:
                    return match.group(0).strip()

            if "LEXIS" in text.upper():
                return text

    return ""


def wait_for_results(driver, wait, old_url):
    try:
        wait.until(
            lambda d:
            d.current_url != old_url
            or len(get_result_cards(d)) > 0
            or "No results" in d.page_source
            or "No Documents Found" in d.page_source
            or "No Results" in d.page_source
        )
    except Exception:
        print("Timed out waiting for results. Continuing anyway.")

    time.sleep(1)


def search_current_source(driver, wait, number, query=None):
    if query is None:
        query = f"number({number})"

    version = get_selected_lexis_version()
    config = LEXIS_CONFIGS[version]

    if version == "plus":
        try:
            toggle_button = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, config["toggle_btn_css"])
                )
            )

            if toggle_button.is_displayed():
                toggle_button.click()
                time.sleep(0.3)

        except Exception:
            pass

    try:
        search_bar = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, SEARCH_BAR_XPATH)
            )
        )

    except Exception:
        mobile_icon = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, ".searchiconmobile")
            )
        )
        mobile_icon.click()

        search_bar = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, SEARCH_BAR_XPATH)
            )
        )

    search_bar.click()
    search_bar.send_keys(Keys.CONTROL, "a")
    search_bar.send_keys(Keys.DELETE)

    old_url = driver.current_url

    search_bar.send_keys(query)
    search_bar.send_keys(Keys.RETURN)

    wait_for_results(driver, wait, old_url)

    time.sleep(0.7)

    results = get_result_cards(driver)

    return results


def get_lexis_data_for_cases(numbers, progress_callback=None):
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless=new")

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()

    links_by_case = {}
    lexis_cites = {}
    hit_counts = {}
    source_used = {}

    zero_hit_numbers = []

    try:
        wait = login_lexis(driver, progress_callback=progress_callback)
        total = len(numbers)

        for index, number in enumerate(numbers, start=1):
            number = str(number).strip()

            if not number:
                links_by_case[number] = ""
                lexis_cites[number] = ""
                hit_counts[number] = 0
                source_used[number] = ""
                continue

            try:
                message = f"Court of Appeal of California cases {index}/{total}: case number {number}"
                print(message)

                if progress_callback:
                    progress_callback(message)

                primary_query = (
                    f"number({number}) and court(Court of Appeal of California)"
                )

                results = search_current_source(
                    driver,
                    wait,
                    number,
                    query=primary_query
                )

                hit_count = len(results)

                hit_counts[number] = hit_count
                source_used[number] = "Court of Appeal of California"

                first_href = ""
                first_cite = ""

                if hit_count >= 1:
                    try:
                        first_href = get_first_href_from_card(results[0])
                    except Exception as e:
                        print(f"Could not get first link for {number}: {type(e).__name__}: {e}")

                if hit_count == 1:
                    try:
                        first_cite = get_lexis_cite_from_card(results[0])
                    except Exception as e:
                        print(f"Could not get Lexis cite for {number}: {type(e).__name__}: {e}")

                if hit_count >= 1:
                    links_by_case[number] = driver.current_url
                else:
                    links_by_case[number] = driver.current_url
                    zero_hit_numbers.append(number)

                lexis_cites[number] = first_cite

                message = f"Court of Appeal of California cases hits for case number {number}: {hit_count}"
                print(message)

                if progress_callback:
                    progress_callback(message)

            except Exception as e:
                message = (
                    f"FAILED Court of Appeal of California cases search "
                    f"for case number {number}: "
                    f"{type(e).__name__}: {e}"
                )
                print(message)

                if progress_callback:
                    progress_callback(message)

                links_by_case[number] = ""
                lexis_cites[number] = ""
                hit_counts[number] = 0
                source_used[number] = "Court of Appeal of California"
                zero_hit_numbers.append(number)

                continue

        if zero_hit_numbers:
            version = get_selected_lexis_version()
            ccc_url = CCC_URLS[version]

            message = f"Starting California Compensation Cases search for {len(zero_hit_numbers)} zero-hit numbers."
            print(message)

            if progress_callback:
                progress_callback(message)

            driver.get(ccc_url)

            wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, SEARCH_BAR_XPATH)
                )
            )

            time.sleep(1)

            fallback_total = len(zero_hit_numbers)

            for index, number in enumerate(zero_hit_numbers, start=1):
                try:
                    message = f"California Compensation Cases fallback search {index}/{fallback_total}: case number {number}"
                    print(message)

                    if progress_callback:
                        progress_callback(message)

                    results = search_current_source(
                        driver,
                        wait,
                        number,
                        query=f"number({number})"
                    )

                    hit_count = len(results)

                    hit_counts[number] = hit_count
                    source_used[number] = "California Compensation Cases"

                    first_href = ""
                    first_cite = ""

                    if hit_count >= 1:
                        try:
                            first_href = get_first_href_from_card(results[0])
                        except Exception as e:
                            print(
                                f"Could not get first CCC link for {number}: "
                                f"{type(e).__name__}: {e}"
                            )

                    if hit_count == 1:
                        try:
                            first_cite = get_lexis_cite_from_card(results[0])
                        except Exception as e:
                            print(
                                f"Could not get CCC Lexis cite for {number}: "
                                f"{type(e).__name__}: {e}"
                            )

                    if hit_count >= 1:
                        links_by_case[number] = driver.current_url

                    if hit_count == 0:
                        lexis_cites[number] = "Missing"
                    else:
                        lexis_cites[number] = first_cite

                    message = f"California Compensation Cases hits for case number {number}: {hit_count}"
                    print(message)

                    if progress_callback:
                        progress_callback(message)

                except Exception as e:
                    message = f"FAILED California Compensation Cases search for case number {number}: {type(e).__name__}: {e}"
                    print(message)

                    if progress_callback:
                        progress_callback(message)

                    try:
                        links_by_case[number] = driver.current_url
                    except Exception:
                        links_by_case[number] = links_by_case.get(number, "")

                    lexis_cites[number] = "Missing"
                    hit_counts[number] = 0
                    source_used[number] = "California Compensation Cases"

                    try:
                        driver.get(ccc_url)

                        wait.until(
                            EC.element_to_be_clickable(
                                (By.XPATH, SEARCH_BAR_XPATH)
                            )
                        )

                        time.sleep(1)

                    except Exception:
                        pass

                    continue

    finally:
        driver.quit()

    return links_by_case, lexis_cites, hit_counts, source_used

def start_export_in_background(preview_window, export_button):
    global current_excel_path

    if not current_excel_path:
        current_excel_path = filedialog.asksaveasfilename(
            title="Save Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="California_HMT_Export.xlsx"
        )

        if not current_excel_path:
            return

    progress = tb.Toplevel(root)
    progress.title("Searching Lexis")
    progress.resizable(True, True)

    center_to_parent(
        progress,
        root,
        600,
        400
    )

    status_var = tk.StringVar(value="Starting Lexis search...")

    status_label = tb.Label(
        progress,
        textvariable=status_var,
        font=("Arial", 11, "bold")
    )
    status_label.pack(pady=(15, 8))

    bar = tb.Progressbar(
        progress,
        mode="indeterminate",
        bootstyle="success-striped",
        length=500
    )
    bar.pack(pady=(0, 10))
    bar.start(10)

    log_frame = tb.Frame(progress)
    log_frame.pack(
        fill=BOTH,
        expand=True,
        padx=15,
        pady=(0, 15)
    )

    log_box = tk.Text(
        log_frame,
        height=16,
        width=70,
        wrap="word",
        state="disabled"
    )

    log_scroll = tb.Scrollbar(
        log_frame,
        orient=VERTICAL,
        command=log_box.yview
    )

    log_box.configure(
        yscrollcommand=log_scroll.set
    )

    log_box.pack(
        side=LEFT,
        fill=BOTH,
        expand=True
    )

    log_scroll.pack(
        side=RIGHT,
        fill=Y
    )

    export_button.config(state="disabled")

    def save_logs_to_file():
        file_path = filedialog.asksaveasfilename(
            title="Save Search Logs",
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt")],
            initialfile=f"California_HMT_Log_{datetime.now():%Y%m%d_%H%M%S}.txt"
        )

        if not file_path:
            return

        try:
            log_text = log_box.get("1.0", "end")

            with open(file_path, "w", encoding="utf-8") as f:
                f.write(log_text)

            show_windows_toast(
                "Logs Saved",
                f"Saved to: {os.path.basename(file_path)}"
            )

        except Exception as e:
            Messagebox.show_error(
                message=str(e),
                title="Save Logs Error",
                parent=progress
            )
    

    def update_progress(message):
        def update_ui():
            status_var.set(message)

            log_box.config(state="normal")
            log_box.insert("end", message + "\n")
            log_box.see("end")
            log_box.config(state="disabled")

        root.after(0, update_ui)
    
    def cleanup():
        try:
            if bar.winfo_exists():
                bar.stop()
        except tk.TclError:
            pass

        try:
            if progress.winfo_exists():
                progress.destroy()
        except tk.TclError:
            pass

    def worker():
        try:
            start_time = time.perf_counter()

            saved_path = export_current_preview_to_excel(
                preview_window=None,
                progress_callback=update_progress
            )

            elapsed = time.perf_counter() - start_time

            minutes = int(elapsed // 60)
            seconds = int(elapsed % 60)

            runtime_text = f"{minutes}m {seconds}s"

            def success_ui():
                status_var.set("Export complete.")
                
                log_box.config(state="normal")
                log_box.insert("end", "Export complete.\n")
                log_box.see("end")
                log_box.config(state="disabled")
                

                try:
                    if preview_window.winfo_exists():
                        preview_window.destroy()
                except tk.TclError:
                    pass

                show_windows_toast(
                "Export Complete",
                f"Saved to: {os.path.basename(saved_path)}\n"
                f"Runtime: {runtime_text}"
                )

                try:
                    should_save_logs = save_logs_var and save_logs_var.get()
                except Exception:
                    should_save_logs = False

                if should_save_logs:
                    save_logs_to_file()

                root.after(1500, cleanup)

            root.after(0, success_ui)

        except Exception as e:
            error_message = str(e)

            def error_ui():
                Messagebox.show_error(
                    message=error_message,
                    title="Export Error"
                )

                cleanup()

            root.after(0, error_ui)



    threading.Thread(target=worker, daemon=True).start()

def show_dataframe_preview(df):
    global current_filtered_df
    global lexis_plus_var
    global save_logs_var

    preview = tb.Toplevel(root)
    preview.title("Data Preview")

    center_to_parent(
        preview,
        root,
        1600,
        700
    )

    main_frame = tb.Frame(preview, padding=10)
    main_frame.pack(fill=BOTH, expand=True)

    status_var = tk.StringVar(
    value=f"Showing 0 of {len(df):,} rows"
    )

    status_frame = tb.Frame(preview)
    status_frame.pack(
        side=BOTTOM,
        fill=X,
        padx=10,
        pady=(0, 5)
    )

    status_label = tb.Label(
        status_frame,
        textvariable=status_var,
        anchor="e",
        font=("Arial", 10)
    )

    status_label.pack(side=RIGHT)

    filter_frame = tb.LabelFrame(
        main_frame,
        text="Filter by Action Type"
    )
    filter_frame.pack(side=LEFT, fill=Y, padx=(0, 10), pady=5)

    table_frame = tb.Frame(main_frame)
    table_frame.pack(side=RIGHT, fill=BOTH, expand=True, padx=(10, 0), pady=5)

    button_frame = tb.Frame(table_frame)
    button_frame.pack(fill=X, pady=(0, 8))

    search_var = tk.StringVar()

    search_entry = tb.Entry(
        button_frame,
        textvariable=search_var,
        width=25
    )
    search_entry.pack(side=LEFT, padx=5)

    search_entry.insert(0, "Search...")

    search_entry.bind(
        "<KeyRelease>",
        lambda event: refresh_preview()
    )

    tree_frame = tb.Frame(table_frame)
    tree_frame.pack(fill=BOTH, expand=True)

    base_df = df.copy()
    excluded_indexes = set()

    action_types = sorted(
        base_df["Action Type"]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )

    action_vars = {}

    canvas = tk.Canvas(
        filter_frame,
        width=300,
        highlightthickness=0
    )

    scrollbar = tb.Scrollbar(
        filter_frame,
        orient=VERTICAL,
        command=canvas.yview,
        # bootstyle="round"
    )

    checkbox_frame = tb.Frame(canvas)

    checkbox_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=checkbox_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side=LEFT, fill=BOTH, expand=True)
    scrollbar.pack(side=RIGHT, fill=Y)

    columns = list(base_df.columns)

    preview_tree = tb.Treeview(
        tree_frame,
        columns=columns,
        show="headings",
        selectmode="extended",
        bootstyle="light"
    )

    y_scroll = tb.Scrollbar(
        tree_frame,
        orient=VERTICAL,
        command=preview_tree.yview
    )

    x_scroll = tb.Scrollbar(
        tree_frame,
        orient=HORIZONTAL,
        command=preview_tree.xview
    )

    preview_tree.configure(
        yscrollcommand=y_scroll.set,
        xscrollcommand=x_scroll.set
    )

    preview_tree.grid(row=0, column=0, sticky=NSEW)
    y_scroll.grid(row=0, column=1, sticky=NS)
    x_scroll.grid(row=1, column=0, sticky=EW)

    tree_frame.rowconfigure(0, weight=1)
    tree_frame.columnconfigure(0, weight=1)

    for col in columns:
        preview_tree.heading(col, text=col, anchor=CENTER)
        preview_tree.column(col, width=180, anchor=CENTER)

    def get_selected_action_types():
        return [
            action_type
            for action_type, var in action_vars.items()
            if var.get()
        ]

    def refresh_preview():
        global current_filtered_df

        for item in preview_tree.get_children():
            preview_tree.delete(item)

        selected_action_types = get_selected_action_types()

        display_df = base_df.drop(
            index=list(excluded_indexes),
            errors="ignore"
        )

        if selected_action_types:
            display_df = display_df[
                display_df["Action Type"]
                .astype(str)
                .str.strip()
                .isin(selected_action_types)
            ]

        search_text = search_var.get().strip().lower()

        if search_text and search_text != "search...":
            display_df = display_df[
                display_df.astype(str)
                .apply(
                    lambda row: row.str.lower().str.contains(
                        search_text,
                        na=False
                    ).any(),
                    axis=1
                )
            ]

        current_filtered_df = display_df.copy()

        for df_index, row in display_df.iterrows():
            preview_tree.insert(
                "",
                END,
                iid=str(df_index),
                values=[row.get(col, "") for col in columns]
            )
        
        status_var.set(
            f"Showing {len(display_df):,} of {len(base_df):,} rows"
        )

    def exclude_selected_rows():
        selected_items = preview_tree.selection()

        if not selected_items:
            Messagebox.show_warning(
                message="Please select one or more rows to exclude.",
                title="No Selection",
                parent=preview
            )
            return

        confirm = Messagebox.yesno(
            message=f"Exclude {len(selected_items)} selected row(s)?",
            title="Exclude Rows",
            parent=preview
        )

        if confirm != "Yes":
            return

        for item in selected_items:
            try:
                excluded_indexes.add(int(item))
            except ValueError:
                pass

        refresh_preview()

    def clear_filters():
        for var in action_vars.values():
            var.set(False)

        refresh_preview()

    def select_all_action_types():
        for var in action_vars.values():
            var.set(True)

        refresh_preview()
    
    def reset_excluded_rows():
        excluded_indexes.clear()
        refresh_preview()

    select_all_button = tb.Button(
        button_frame,
        text="Select All Action Types",
        command=select_all_action_types,
        bootstyle="info"
    )
    select_all_button.pack(side=LEFT, padx=5)

    clear_button = tb.Button(
        button_frame,
        text="Clear Action Type Filters",
        command=clear_filters,
        bootstyle="primary"
    )
    clear_button.pack(side=LEFT, padx=5)

    exclude_button = tb.Button(
        button_frame,
        text="Exclude Selected Row(s)",
        command=exclude_selected_rows,
        bootstyle="secondary"
    )
    exclude_button.pack(side=LEFT, padx=5)

    reset_button = tb.Button(
        button_frame,
        text="Reset Filter",
        command=reset_excluded_rows,
        bootstyle="warning"
    )
    reset_button.pack(side=LEFT, padx=5)

    view_pdf_button = tb.Button(
        button_frame,
        text="View PDF",
        command=open_downloaded_pdf,
        bootstyle="danger"
    )
    view_pdf_button.pack(side=LEFT, padx=5)

    export_button = tb.Button(
        button_frame,
        text="Search and Export",
        command=lambda: start_export_in_background(preview, export_button),
        bootstyle="success"
    )
    export_button.pack(side=LEFT, padx=5)

    lexis_plus_var = tk.BooleanVar(value=False)

    lexis_label_var = tk.StringVar(
        value="Lexis Advance"
    )

    def update_lexis_label():
        if lexis_plus_var.get():
            lexis_label_var.set("Lexis Plus")
        else:
            lexis_label_var.set("Lexis Advance")

    lexis_toggle = tb.Checkbutton(
        button_frame,
        
        variable=lexis_plus_var,
        command=update_lexis_label,
        bootstyle="danger-round-toggle"
    )
    lexis_toggle.pack(side=LEFT, padx=5)

    lexis_label = tb.Label(
        button_frame,
        textvariable=lexis_label_var
    )
    lexis_label.pack(side=LEFT, padx=5)

    save_logs_var = tk.BooleanVar(value=False)

    save_logs_check = tb.Checkbutton(
        button_frame,
        text="Save Logs",
        variable=save_logs_var,
        bootstyle="round-toggle"
    )
    save_logs_check.pack(side=LEFT, padx=5)

    for i, action_type in enumerate(action_types):
        var = tk.BooleanVar(value=False)
        action_vars[action_type] = var

        checkbox = tb.Checkbutton(
            checkbox_frame,
            text=action_type,
            variable=var,
            command=refresh_preview,
            # bootstyle="round-toggle",
            width=35
        )

        checkbox.grid(row=i, column=0, sticky=W, padx=5, pady=3)

    refresh_preview()


def load_links():
    for item in tree.get_children():
        tree.delete(item)

    try:
        links = get_conference_links()

        if not links:
            Messagebox.show_warning(
                "No conference links were found.",
                "No Links Found"
            )
            return

        for date_text, href in links:
            tree.insert(
                "",
                END,
                values=(date_text, href)
            )

        show_windows_toast(
            "Links Loaded",
            f"Loaded {len(links)} conference link(s)."
        )

    except Exception as e:
        Messagebox.show_error(str(e), "Error")

def import_pdf_file():
    global downloaded_pdf_path
    global current_excel_path
    global current_filtered_df
    global imported_pdf_file
    imported_pdf_file = True

    file_path = filedialog.askopenfilename(
        title="Select PDF File",
        filetypes=[("PDF Files", "*.pdf")]
    )

    if not file_path:
        return

    downloaded_pdf_path = file_path
    current_filtered_df = None

    safe_name = Path(file_path).stem

    current_excel_path = None

    convert_preview_downloaded_pdf()

def download_selected_pdf():
    global downloaded_pdf_path
    global current_excel_path
    global current_filtered_df
    global imported_pdf_file
    imported_pdf_file = False
    global current_conference_date

    

    selected = tree.selection()

    if not selected:
        Messagebox.show_warning(
            "Please select a link first.",
            "No Selection"
        )
        return

    values = tree.item(selected[0], "values")

    date_text = values[0]
    url = values[1]

    current_conference_date = date_text

    try:
        response = requests.get(url, timeout=30)

        if response.status_code != 200:
            Messagebox.show_error(
                f"HTTP Status: {response.status_code}",
                "Download Failed"
            )
            return

        safe_date = (
            date_text
            .replace(",", "")
            .replace(" ", "_")
            .replace("/", "-")
        )

        current_excel_path = None

        if downloaded_pdf_path and os.path.exists(downloaded_pdf_path):
            os.remove(downloaded_pdf_path)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
            temp_pdf.write(response.content)
            downloaded_pdf_path = temp_pdf.name

        current_filtered_df = None

    except Exception as e:
        Messagebox.show_error(str(e), "Error")


def download_and_preview_selected_pdf():
    download_selected_pdf()

    if downloaded_pdf_path and os.path.exists(downloaded_pdf_path):
        convert_preview_downloaded_pdf()


def convert_preview_downloaded_pdf():
    global current_filtered_df

    if not downloaded_pdf_path or not os.path.exists(downloaded_pdf_path):
        Messagebox.show_warning(
            "Please download a PDF first.",
            "No PDF"
        )
        return

    try:
        df = extract_table_by_positions(downloaded_pdf_path)

        if df.empty:
            Messagebox.show_warning(
                "No rows were extracted from the PDF.",
                "No Data"
            )
            return

        current_filtered_df = df.copy()
        show_dataframe_preview(df)

    except Exception as e:
        Messagebox.show_error(str(e), "Error")


def format_excel_file(excel_path, header_row=5, last_row=None):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    ws = wb.active

    if last_row is None:
        last_row = ws.max_row

    # Freeze header row
    # ws.freeze_panes = "A2"

    # Add filters
    # ws.auto_filter.ref = ws.dimensions

    # Fills
    
    header_fill = PatternFill(
        fill_type="solid",
        start_color="F46036",
        end_color="F46036"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFEE93",
        end_color="FFEE93"
    )

    gray_fill = PatternFill(
        fill_type="solid",
        start_color="D7D6D6",
        end_color="D7D6D6"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FF5A5F",
        end_color="FF5A5F"
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAD3",
        end_color="D9EAD3"
    )

    # Borders
    thin = Side(
        border_style="thin",
        color="D9D9D9"
    )

    medium = Side(
        border_style="medium",
        color="000000"
    )

    normal_border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    header_border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=medium
    )

    # Header formatting
    HEADER_ROW = header_row

    first_info_row = HEADER_ROW - 4
    last_info_row = HEADER_ROW - 2

    for row in (HEADER_ROW - 4, HEADER_ROW - 3, HEADER_ROW - 2):
        cell = ws[f"A{row}"]

        cell.fill = header_fill

        cell.font = Font(
            bold=True,
            color="FFFFFF",
            size=12
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        cell.border = Border(
            left=medium,
            right=medium,
            top=medium if row == first_info_row else thin,
            bottom=medium if row == last_info_row else thin
        )

    for cell in ws[HEADER_ROW]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = header_border

    # Apply borders and alignment to all cells
    for row in ws.iter_rows(
        min_row=HEADER_ROW + 1,
        max_row=last_row
    ):
        for cell in row:
            cell.border = normal_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # Alternate row shading
    for row_num in range(
        HEADER_ROW + 1,
        last_row + 1
    ):
        if row_num % 2 == 0:
            for cell in ws[row_num]:
                cell.fill = gray_fill

    # Find important columns
    lexis_cite_col = None
    links_col = None

    HEADER_ROW = header_row

    for cell in ws[HEADER_ROW]:
        if cell.value == "Lexis Cite":
            lexis_cite_col = cell.column
        elif cell.value == "Links":
            links_col = cell.column

    # Highlight blank Lexis Cite cells
    if lexis_cite_col:
        for row in range(HEADER_ROW + 1, last_row + 1):
            cell = ws.cell(row=row, column=lexis_cite_col)

            value = "" if cell.value is None else str(cell.value).strip()

            if value == "":
                cell.fill = red_fill

            elif value == "Missing":
                cell.fill = yellow_fill

            else:
                cell.fill = green_fill

    # Make links clickable and cleaner
    if links_col:
        for row in range(HEADER_ROW + 1, last_row + 1):
            cell = ws.cell(row=row, column=links_col)

            if cell.value:
                url = str(cell.value).strip()

                if url.lower().startswith("http"):
                    cell.value = "Open Link"
                    cell.hyperlink = url
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
                else:
                    cell.value = url

    # Auto-size columns with width cap
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(excel_path)


def export_current_preview_to_excel(preview_window=None, progress_callback=None):
    global current_filtered_df
    global current_excel_path
    global downloaded_pdf_path
    global imported_pdf_file

    if current_filtered_df is None or current_filtered_df.empty:
        raise RuntimeError("Please convert and preview the PDF first.")

    if not downloaded_pdf_path:
        raise RuntimeError("Please download or import a PDF first.")

    df = current_filtered_df.copy()

    numbers = (
        df["CA #"]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )

    print("CA numbers to search:", len(numbers))

    (
        links_by_case,
        lexis_cites,
        hit_counts,
        source_used
    ) = get_lexis_data_for_cases(
        numbers,
        progress_callback=progress_callback
    )

    number_key = df["CA #"].astype(str).str.strip()

    df["Lexis Cite"] = number_key.map(lexis_cites)
    df["Questions/Issues Encountered"] = ""
    df["Email"] = ""
    df["Hits"] = number_key.map(hit_counts)
    df["Source"] = number_key.map(source_used)
    df["Links"] = number_key.map(links_by_case)

    file_exists = os.path.exists(current_excel_path)

    if file_exists:
        wb = load_workbook(current_excel_path)
        ws = wb.active

        startrow = ws.max_row + 4

        ws[f"A{startrow + 1}"] = os.environ.get("USERNAME", "")
        ws[f"A{startrow + 2}"] = (
            f"Conference Results for {current_conference_date}"
        )
        ws[f"A{startrow + 3}"] = (
            "Last Run: "
            + datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
        )

        wb.save(current_excel_path)

        with pd.ExcelWriter(
            current_excel_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay"
        ) as writer:
            df.to_excel(
                writer,
                index=False,
                header=True,
                startrow=startrow + 4
            )
        header_row = startrow + 5
        first_data_row = header_row + 1
        last_row = first_data_row + len(df) - 1

        format_excel_file(
            current_excel_path,
            header_row=header_row,
            last_row=last_row
        )

    else:
        with pd.ExcelWriter(
            current_excel_path,
            engine="openpyxl"
        ) as writer:
            df.to_excel(
                writer,
                index=False,
                startrow=4
            )

            ws = writer.book.active

            ws["A1"] = os.environ.get("USERNAME", "")
            ws["A2"] = f"Conference Results for {current_conference_date}"
            ws["A3"] = (
                "Last Run: "
                + datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
            )

        header_row = 5
        first_data_row = 6
        last_row = first_data_row + len(df) - 1

        format_excel_file(
            current_excel_path,
            header_row=header_row,
            last_row=last_row
        )

    if (
        downloaded_pdf_path
        and os.path.exists(downloaded_pdf_path)
        and not imported_pdf_file
    ):
        os.remove(downloaded_pdf_path)
        downloaded_pdf_path = None

    return current_excel_path


print("Starting app...")

root = tb.Window(themename="superhero")
root.title("California HMT Tool")

width = 600
height = 500

root.update_idletasks()

x = (root.winfo_screenwidth() // 2) - (width // 2)
y = (root.winfo_screenheight() // 2) - (height // 2)

root.geometry(f"{width}x{height}+{x}+{y}")

main_container = tb.Frame(root, padding=15)
main_container.pack(fill=BOTH, expand=True)

title = tb.Label(
    main_container,
    text="Weekly Conference Results",
    font=("Arial", 20, "bold")
)
title.pack(pady=(0, 12))

style = tb.Style()

theme_frame = tb.Frame(main_container)
theme_frame.pack(pady=(0, 10))

theme_label = tb.Label(
    theme_frame,
    text="Theme:"
)
theme_label.pack(side=LEFT, padx=(0, 5))

theme_combo = tb.Combobox(
    theme_frame,
    values=style.theme_names(),
    state="readonly",
    width=18
)
theme_combo.set(style.theme_use())
theme_combo.pack(side=LEFT)

def change_theme(event=None):
    selected_theme = theme_combo.get()
    style.theme_use(selected_theme)

theme_combo.bind("<<ComboboxSelected>>", change_theme)

button_frame = tb.Frame(main_container)
button_frame.pack(pady=(0, 10))

button_frame = tb.Frame(main_container)
button_frame.pack(pady=(0, 10))

conference_button = tb.Button(
    button_frame,
    text="Weekly Conference Results Site",
    command=open_conference_website,
    bootstyle="info"
)
conference_button.pack(side=LEFT, padx=5)

import_pdf_button = tb.Button(
    button_frame,
    text="Import PDF",
    command=import_pdf_file,
    bootstyle="secondary"
)
import_pdf_button.pack(side=LEFT, padx=5)

load_button = tb.Button(
    button_frame,
    text="Load Links",
    command=start_load_links_in_background,
    bootstyle="success"
)
load_button.pack(side=LEFT, padx=5)

download_preview_button = tb.Button(
    button_frame,
    text="Download and Preview",
    command=download_and_preview_selected_pdf,
    bootstyle="danger"
)
download_preview_button.pack(side=LEFT, padx=5)

tree_frame = tb.Frame(main_container)
tree_frame.pack(fill=BOTH, expand=True)

columns = ("Date", "Href")

tree = tb.Treeview(
    tree_frame,
    columns=columns,
    show="headings",
    bootstyle="light"
)

tree_scroll = tb.Scrollbar(
    tree_frame,
    orient=VERTICAL,
    command=tree.yview
)

tree.configure(yscrollcommand=tree_scroll.set)

tree.heading("Date", text="Conference Dates", anchor=CENTER)
tree.heading("Href", text="Href", anchor=CENTER)

tree.column("Date", width=250, anchor=CENTER)
tree.column("Href", width=0, stretch=False)

tree.grid(row=0, column=0, sticky=NSEW)
tree_scroll.grid(row=0, column=1, sticky=NS)

tree_frame.rowconfigure(0, weight=1)
tree_frame.columnconfigure(0, weight=1)

root.mainloop()